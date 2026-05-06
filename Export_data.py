def export_to_excel(parent, optimization_data, financial_data, emissions_data=None, metadata_dict=None, selected_inputs=None, tab_name="optimization"):
    try:
        import numpy as np
        import pandas as pd
        import openpyxl
        from openpyxl.utils import get_column_letter
        from PyQt5.QtWidgets import QFileDialog, QMessageBox


        def _to_len(a, n):
            arr = np.asarray(a).flatten()
            return np.resize(arr, n)


        def _write_stochastic_block(writer, sheet_name, existing_cols_count, index_like, series_tuple, label_prefix, max_len):
            if not series_tuple:
                return
            data = {}
            for i, s in enumerate(series_tuple, start=1):
                try:
                    arr = np.zeros(max_len) if s is None else _to_len(s, max_len)
                except Exception:
                    arr = np.zeros(max_len)
                data[f"{label_prefix} (S{i})"] = arr
            df_stoch = pd.DataFrame(data, index=index_like)
            startcol = 1 + existing_cols_count   
            df_stoch.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=startcol)

        def _canonical_sheet_name(raw_key: str) -> str:
            base = (raw_key or "")
            base = base.replace("Input:", "").replace("Data:", "")
            base = base.replace("Inputs", "").replace(" Data", "").replace(" + ", " and ")
            base = base.strip()
            lower = base.lower()

            if any(x in lower for x in ["stochastic", "stoch"]):
                return "__SKIP__"
            if "co₂" in lower or "co2" in lower:
                return "__SKIP__"
            if "external cost" in lower:
                return "__SKIP__"

            if "grid power limit" in lower or lower == "grid":
                return "__SKIP__"

            if "price data inputs" in lower:
                return "__SKIP__"

            if "pv" in lower:
                return "PV Generation"
            if "wind" in lower:
                return "Wind Generation"
            if ("solar" in lower and "collector" in lower) or "solar thermal" in lower or "collectora" in lower:
                return "Solar Collector"
            if "thermal and electricity demand" in lower:
                return "Thermal and Electricity Demand"
            if "electricity demand" in lower:
                return "Electricity Demand"
            if ("thermal" in lower and "demand" in lower) or ("heating" in lower and "demand" in lower) or ("cooling" in lower and "demand" in lower):
                return "Thermal Demand"
            if "price" in lower:
                return "Price"
            return base

        def _get_first_present(d, *names):
            if not isinstance(d, dict):
                return None
            for n in names:
                if n in d:
                    return d[n]
            return None

        def _extract_emission_sections(emissions_data):
            el_dict = None
            th_dict = None
            if isinstance(emissions_data, dict):
                el_dict = emissions_data.get("el_emission_dict") or emissions_data.get("electricity") or emissions_data.get("Electricity")
                th_dict = emissions_data.get("th_emission_dict") or emissions_data.get("thermal") or emissions_data.get("Thermal")
                if not el_dict or not isinstance(el_dict, dict):
                    for k, v in emissions_data.items():
                        if isinstance(v, dict) and "electric" in str(k).lower():
                            el_dict = v; break
                if not th_dict or not isinstance(th_dict, dict):
                    for k, v in emissions_data.items():
                        if isinstance(v, dict) and "therm" in str(k).lower():
                            th_dict = v; break
                if not el_dict or not isinstance(el_dict, dict):
                    for v in emissions_data.values():
                        if isinstance(v, dict) and any("co2" in k.lower() or "co₂" in k.lower() for k in v.keys()):
                            el_dict = v; break
                if not th_dict or not isinstance(th_dict, dict):
                    for v in emissions_data.values():
                        if v is not el_dict and isinstance(v, dict) and any("co2" in k.lower() or "co₂" in k.lower() for k in v.keys()):
                            th_dict = v; break
            return el_dict if isinstance(el_dict, dict) else None, th_dict if isinstance(th_dict, dict) else None

        def _build_co2_df(el_dict, th_dict):
            rows = []
            def push(section_name, d):
                if not isinstance(d, dict):
                    return
                rows.append([section_name, "The following data refers only to", d.get("\nThe following data refers only to", section_name)])
                for k, v in d.items():
                    key_l = k.lower()
                    if "co₂" in key_l or "co2" in key_l:
                        rows.append([section_name, k, v])
            push("Electricity", el_dict)
            push("Thermal", th_dict)
            return pd.DataFrame(rows, columns=["Section", "Metric", "Value"]) if rows else None

        def _build_external_cost_df(el_dict, th_dict):
            rows = []
            def push(section_name, d):
                if not isinstance(d, dict):
                    return
                rows.append([section_name, "The following data refers only to", d.get("\nThe following data refers only to", section_name)])
                for k, v in d.items():
                    if "external" in k.lower():
                        rows.append([section_name, k, v])
            push("Electricity", el_dict)
            push("Thermal", th_dict)
            return pd.DataFrame(rows, columns=["Section", "Metric", "Value"]) if rows else None

        # Information sheet helpers
        def _info_mapping():
            """Units and explanations (EN) for known variables."""
            kW = "kW (hourly average)"
            kW_th = "kW_th (hourly average)"
            kW_cool = "kW_cool (hourly average)"
            kWh = "kWh"
            EURkWh = "€/kWh"
            EURkWh_th = "€/kWh_th"
            EUR_t = "€ per time step"

            return {
                # Generation
                "PV Generation": (kW, "Electrical power produced by the PV system. Integrates over time to energy [kWh]."),
                "Wind Generation": (kW, "Electrical power produced by the wind turbine. Integrates to [kWh]."),
                "Solar Collector Generation": (kW_th, "Thermal output from solar thermal collectors. Integrates to [kWh_th]."),

                # Energy flows
                "PV → Load": (kW, "PV electrical power supplied directly to the electrical load."),
                "PV → Grid": (kW, "PV electrical power exported to the grid."),
                "PV → Battery": (kW, "PV electrical power used to charge the battery."),
                "PV → Heat Pump (heating)": (kW, "PV electrical power to the heat pump compressor (heating mode)."),
                "PV → Heat Pump (cooling)": (kW, "PV electrical power to the heat pump compressor (cooling mode)."),

                "Wind → Load": (kW, "Wind electrical power supplied to the electrical load."),
                "Wind → Grid": (kW, "Wind electrical power exported to the grid."),
                "Wind → Battery": (kW, "Wind electrical power used to charge the battery."),
                "Wind → Heat Pump (heating)": (kW, "Wind electrical power to the heat pump compressor (heating mode)."),
                "Wind → Heat Pump (cooling)": (kW, "Wind electrical power to the heat pump compressor (cooling mode)."),

                "Grid → Load": (kW, "Imported grid power to supply the electrical load."),
                "Grid → Battery": (kW, "Imported grid power used to charge the battery."),
                "Grid → Heat Pump (heating)": (kW, "Imported grid power to the heat pump compressor (heating mode)."),
                "Grid → Heat Pump (cooling)": (kW, "Imported grid power to the heat pump compressor (cooling mode)."),

                "Battery → Load": (kW, "Battery discharge power to the electrical load."),
                "Battery → Grid": (kW, "Battery discharge power exported to the grid."),
                "Battery → Heat Pump (heating)": (kW, "Battery discharge power to the heat pump compressor (heating mode)."),
                "Battery → Heat Pump (cooling)": (kW, "Battery discharge power to the heat pump compressor (cooling mode)."),
                "Battery charge": (kW, "Total electrical charging power into the battery (from all sources)."),
                "Battery discharge": (kW, "Electrical discharging power from the battery."),
                "Battery SOE": (kWh, "Battery state of energy."),

                # Heat pump, buffer tank
                "Heat Pump → Heating Load + Buffer Tank": (kW_th, "Heat pump thermal output to space heating and/or buffer tank."),
                "Heat Pump → Heating Load": (kW_th, "Heat pump thermal output to space heating."),
                "Heat Pump → Cooling Load": (kW_cool, "Heat pump cooling capacity delivered to cooling load."),
                "Heat Pump → Buffer Tank": (kW_th, "Heat pump thermal output to the buffer tank."),

                "Buffer Tank → Heating Load": (kW_th, "Thermal discharge from buffer tank to the heating load."),
                "Buffer Tank SOE": (kWh, "Thermal state of energy of the buffer tank."),

                "Solar Collector → Buffer Tank": (kW_th, "Thermal flow from solar collectors into the buffer tank."),
                "Unmet Solar Collector → Buffer Tank": (kW_th, "Curtailment of potential solar collector heat to the tank due to system limits."),

                # Demands
                "Electricity Demand": (kW, "Electrical demand of the system."),
                "Heating Demand": (kW_th, "Thermal demand for space heating."),
                "Cooling Demand": (kW_cool, "Cooling demand."),

                # Unmet demand & curtailments
                "Unmet Electricity Demand": (kW, "Portion of electrical demand that remains unmet."),
                "Unmet Heating Demand": (kW_th, "Portion of heating demand that remains unmet."),
                "Unmet Cooling Demand": (kW_cool, "Portion of cooling demand that remains unmet."),
                "PV Lost": (kW, "Curtailed/unused PV power."),
                "Wind Lost": (kW, "Curtailed/unused wind power."),

                # Prices & finance
                "Buy Price": (EURkWh, "Electricity import price from the grid."),
                "Sell Price": (EURkWh, "Electricity export (feed-in) price to the grid."),
                "Thermal Price": (EURkWh_th, "Valuation or equivalent price of thermal energy."),
                "Revenue": (EUR_t, "Revenue from electricity exported to the grid (per time step)."),
                "Cost": (EUR_t, "Cost of imported electricity from the grid (per time step)."),
                "Savings": (EUR_t, "Avoided cost from self-consumption of electricity and/or thermal energy (per time step)."),
                "Net profit": (EUR_t, "Revenue minus Cost (per time step)."),
            }

        def _infer_unit(name: str) -> str:
            """Heuristic for unseen variables."""
            n = (name or "").lower()
            if "price" in n:
                return "€/kWh" if "thermal" not in n else "€/kWh_th"
            if "soe" in n:
                return "kWh" if "buffer" not in n else "kWh_th"
            if "heating" in n or "thermal" in n or "collector" in n or "buffer" in n:
                return "kW_th (hourly average)"
            if "cool" in n:
                return "kW_cool (hourly average)"
            if any(x in n for x in ["revenue", "cost", "saving", "profit"]):
                return "€ per time step"
            return "kW (hourly average)"

        def _default_description(name: str) -> str:
            if "→" in name:
                src, dst = [s.strip() for s in name.split("→", 1)]
                return f"Energy flow from {src} to {dst}."
            if "demand" in name.lower():
                return "System demand."
            if "price" in name.lower():
                return "Energy price per unit."
            if "soe" in name.lower():
                return "State of energy in the storage."
            return "Model variable (auto-generated description)."

        def _build_information_df(opt_dict: dict) -> pd.DataFrame:
            mapping = _info_mapping()
            rows = []
            for key in sorted(opt_dict.keys(), key=str):
                unit, desc = mapping.get(key, (_infer_unit(key), _default_description(key)))
                rows.append((key, unit, desc))
            return pd.DataFrame(rows, columns=["Variable", "Unit", "Explanation"])


        max_len = 8760
        default_filename = f"{tab_name}.xlsx"
        path, _ = QFileDialog.getSaveFileName(parent, "Save Excel File", default_filename, "Excel Files (*.xlsx)")
        if not path:
            return


        stoch = None
        if isinstance(selected_inputs, dict):
            stoch = selected_inputs.get("stochastic")

        with pd.ExcelWriter(path, engine="openpyxl") as writer:

            df_opt = pd.DataFrame({
                k: np.resize(np.asarray(v).flatten(), max_len)
                for k, v in optimization_data.items()
                if v is not None and k != "metadata"
            }).round(4)
            df_opt.index.name = "Hour"
            df_opt.to_excel(writer, sheet_name="Optimization", index=True)


            try:
                df_info = _build_information_df({k: v for k, v in optimization_data.items() if k != "metadata"})
                df_info.to_excel(writer, sheet_name="Information", index=False)
            except Exception:
                pd.DataFrame([{"Variable": "Error", "Unit": "-", "Explanation": "Failed to create the information list."}]).to_excel(
                    writer, sheet_name="Information", index=False
                )


            if isinstance(financial_data, dict):
                df_fin = pd.DataFrame(financial_data.items(), columns=["Metric", "Value"])
            
                old_series = df_fin.loc[df_fin["Metric"] == "_old_thermal_cost", "Value"]
                old_val = pd.to_numeric(old_series.astype(str).str.replace(",", ".", regex=False), errors="coerce").fillna(0.0)
                old_val = float(old_val.iloc[0]) if len(old_val) else 0.0
            
                if abs(old_val) < 1e-12:
                    df_fin = df_fin[~df_fin["Metric"].isin(["_old_thermal_cost", "_new_thermal_cost", "_diff_thermal_cost"])]
                
                rename_map = {
                    "_old_thermal_cost": "Old heating cost [€]",
                    "_new_thermal_cost": "New heating cost [€]",
                    "_diff_thermal_cost": "Difference in old and new heating cost [€]"
                }
                
                df_fin["Metric"] = df_fin["Metric"].replace(rename_map)
                
                
                df_fin.to_excel(writer, sheet_name="Financial Summary", index=False)
            else:
                pd.DataFrame([{"Metric": "Export Error", "Value": "Financial data is not a dictionary"}]).to_excel(
                    writer, sheet_name="Financial Summary", index=False
                )



            if emissions_data:
                if isinstance(emissions_data, dict):
                    pd.DataFrame(emissions_data.items(), columns=["Metric", "Value"]).to_excel(
                        writer, sheet_name="Emissions Summary", index=False
                    )
                else:
                    pd.DataFrame([{"Metric": "Export Error", "Value": "Emissions data is not a dictionary"}]).to_excel(
                        writer, sheet_name="Emissions Summary", index=False
                    )


            el_d, th_d = _extract_emission_sections(emissions_data)
            df_co2 = _build_co2_df(el_d, th_d)
            if df_co2 is not None and not df_co2.empty:
                df_co2.to_excel(writer, sheet_name="CO₂ Emissions", index=False)
            df_ext = _build_external_cost_df(el_d, th_d)
            if df_ext is not None and not df_ext.empty:
                df_ext.to_excel(writer, sheet_name="External Cost", index=False)


            processed_categories = set()
            keys_iter = (selected_inputs.keys() if isinstance(selected_inputs, dict) else (selected_inputs or []))
            for raw_key in keys_iter:
                canon = _canonical_sheet_name(str(raw_key))
                if canon in ("__SKIP__", ""):
                    continue
                if canon in processed_categories:
                    continue
                processed_categories.add(canon)


                if canon == "PV Generation":
                    relevant_cols = [k for k in df_opt.columns if k.startswith("PV →") or k in ("PV Generation", "PV Lost")]
                elif canon == "Wind Generation":
                    relevant_cols = [k for k in df_opt.columns if k.startswith("Wind →") or k in ("Wind Generation", "Wind Lost")]
                elif canon == "Electricity Demand":
                    relevant_cols = [k for k in df_opt.columns if "Electricity Demand" in k]
                elif canon == "Thermal Demand":
                    relevant_cols = [k for k in df_opt.columns if ("Heating Demand" in k or "Cooling Demand" in k)]
                elif canon == "Thermal and Electricity Demand":
                    relevant_cols = [k for k in df_opt.columns if ("Electricity Demand" in k or "Heating Demand" in k or "Cooling Demand" in k)]
                elif canon == "Price":
                    relevant_cols = [c for c in ("Buy Price", "Sell Price") if c in df_opt.columns]
                else:
                    relevant_cols = [k for k in df_opt.columns if canon in k]


                if relevant_cols:
                    df_subset = df_opt[relevant_cols].round(4)
                    df_subset.index.name = "Hour"
                    df_subset.to_excel(writer, sheet_name=canon, index=True)
                    deterministic_index = df_subset.index
                    existing_cols_count = len(relevant_cols)
                else:
                    empty_index = pd.RangeIndex(start=0, stop=max_len, step=1)
                    df_idx = pd.DataFrame(index=empty_index)
                    df_idx.index.name = "Hour"
                    df_idx.to_excel(writer, sheet_name=canon, index=True)
                    deterministic_index = df_idx.index
                    existing_cols_count = 0


                if isinstance(stoch, dict):
                    if canon == "PV Generation":
                        series = _get_first_present(stoch, "PV Generation", "PV", "Data: PV Generation")
                        if series is not None:
                            series_tuple = tuple(series) if isinstance(series, (list, tuple)) else (series,)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, series_tuple, "PV", max_len)

                    elif canon == "Wind Generation":
                        series = _get_first_present(stoch, "Wind Generation", "Wind", "Data: Wind Generation")
                        if series is not None:
                            series_tuple = tuple(series) if isinstance(series, (list, tuple)) else (series,)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, series_tuple, "Wind", max_len)

                    elif canon == "Solar Collector":
                        series = _get_first_present(stoch, "Solar Collector Generation", "Solar Collector", "Solar Thermal", "Data: Solar Collector Generation")
                        if series is not None:
                            series_tuple = tuple(series) if isinstance(series, (list, tuple)) else (series,)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, series_tuple, "Solar Collector", max_len)

                    elif canon == "Electricity Demand":
                        series = _get_first_present(stoch, "Electricity Demand", "Electric Demand", "Elec Demand", "Data: Electricity Demand")
                        if series is not None:
                            series_tuple = tuple(series) if isinstance(series, (list, tuple)) else (series,)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, series_tuple, "Electricity Demand", max_len)

                    elif canon == "Price":
                        series = _get_first_present(stoch, "Price Data", "Price", "Buy Price", "Data: Price Data")
                        if series is not None:
                            series_tuple = tuple(series) if isinstance(series, (list, tuple)) else (series,)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, series_tuple, "Price", max_len)

                    elif canon == "Thermal Demand":
                        td = _get_first_present(stoch, "Thermal Demand", "Heat/Cool Demand", "Data: Thermal Demand")
                        if td:
                            heating_series = []
                            cooling_series = []
                            if isinstance(td, dict):
                                h_list = td.get("heating") or []
                                c_list = td.get("cooling") or []
                                if not isinstance(h_list, (list, tuple)): h_list = [h_list]
                                if not isinstance(c_list, (list, tuple)): c_list = [c_list]
                                heating_series.extend(h_list)
                                cooling_series.extend(c_list)
                            elif isinstance(td, (list, tuple)):
                                for item in td:
                                    if isinstance(item, dict):
                                        heating_series.append(item.get("heating"))
                                        cooling_series.append(item.get("cooling"))
                                    else:
                                        heating_series.append(None)
                                        cooling_series.append(None)

                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, tuple(heating_series), "Heating", max_len)
                            existing_cols_count += len(heating_series)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, tuple(cooling_series), "Cooling", max_len)

                    elif canon == "Thermal and Electricity Demand":
                        series = _get_first_present(stoch, "Electricity Demand", "Electric Demand", "Elec Demand", "Data: Electricity Demand")
                        if series is not None:
                            series_tuple = tuple(series) if isinstance(series, (list, tuple)) else (series,)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, series_tuple, "Electricity Demand", max_len)
                            existing_cols_count += len(series_tuple)

                        td = _get_first_present(stoch, "Thermal Demand", "Heat/Cool Demand", "Data: Thermal Demand")
                        if td:
                            heating_series = []
                            cooling_series = []
                            if isinstance(td, dict):
                                h_list = td.get("heating") or []
                                c_list = td.get("cooling") or []
                                if not isinstance(h_list, (list, tuple)): h_list = [h_list]
                                if not isinstance(c_list, (list, tuple)): c_list = [c_list]
                                heating_series.extend(h_list)
                                cooling_series.extend(c_list)
                            elif isinstance(td, (list, tuple)):
                                for item in td:
                                    if isinstance(item, dict):
                                        heating_series.append(item.get("heating"))
                                        cooling_series.append(item.get("cooling"))
                                    else:
                                        heating_series.append(None)
                                        cooling_series.append(None)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, tuple(heating_series), "Heating", max_len)
                            existing_cols_count += len(heating_series)
                            _write_stochastic_block(writer, canon, existing_cols_count, deterministic_index, tuple(cooling_series), "Cooling", max_len)


            input_records = []
            if isinstance(metadata_dict, dict):
                for tech_name, params in metadata_dict.items():
                    if isinstance(params, dict):
                        for param, val in params.items():
                            input_records.append((tech_name, param, val))

            if isinstance(selected_inputs, dict):
                if "Grid Power Limit" in selected_inputs:
                    input_records.append(("Grid", "Grid Power Limit", selected_inputs["Grid Power Limit"]))
                if ("stochastic" in selected_inputs):
                    input_records.append(("Model", "Type", "two_stage_stochastic"))

            if input_records:
                pd.DataFrame(input_records, columns=["Input Category", "Parameter", "Value"]).to_excel(
                    writer, sheet_name="INPUT", index=False
                )


        wb = openpyxl.load_workbook(path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.freeze_panes = "B2" if sheet != "Financial Summary" else "A2"
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                header = column_cells[0].value
                header_len = len(str(header)) if header is not None else 0
                cell_lens = [len(str(c.value)) for c in column_cells[1:] if c.value is not None]
                avg_len = (sum(cell_lens) / len(cell_lens)) if cell_lens else 0
                ws.column_dimensions[get_column_letter(col_idx)].width = max(avg_len, header_len) + 2
        wb.save(path)

        QMessageBox.information(parent, "Export Complete", f"Results exported to:\n{path}")

    except Exception as e:
        QMessageBox.critical(parent, "Export Failed", f"Error:\n{str(e)}")
